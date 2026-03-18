import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Sample data
data = {
    "Name": ["Amit", "Priya", "Rahul"],
    "Score": [85, 90, 78]
}

df = pd.DataFrame(data)
file_name = "students_scores.xlsx"

# Save DataFrame to Excel
df.to_excel(file_name, index=False)

# Load workbook with openpyxl
wb = load_workbook(file_name)
ws = wb.active

# Make header bold
for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save(file_name)
print("Excel file created with bold headers!")



